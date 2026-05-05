import cloudscraper
import re
import feedparser
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# ================= CONFIG =================

RSS_URL = "https://www.bleepingcomputer.com/feed/"
CVE_PATTERN = r"CVE[\s\-–—]*\d{4}[\s\-–—]*\d{4,7}"
FILE = "data.xlsx"

scraper = cloudscraper.create_scraper()

# ================= HELPERS =================

def clean_cve(cve):
    return re.sub(r"[\s–—\-]+", "-", cve.upper()).strip("-")

def extract_cves(url):
    try:
        res = scraper.get(url, timeout=10)
    except:
        return set()

    soup = BeautifulSoup(res.text, "html.parser")
    text = soup.get_text()

    raw = re.findall(CVE_PATTERN, text, re.IGNORECASE)
    return {clean_cve(c) for c in raw}

def load_existing():
    if not os.path.exists(FILE):
        return set()

    wb = load_workbook(FILE)
    ws = wb.active

    # Only CVE used for uniqueness
    return {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}

def save_new(data):
    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["CVE", "DATE", "LINK"])   # 🔥 3 columns
    else:
        wb = load_workbook(FILE)
        ws = wb.active

    for cve, date, link in data:
        ws.append([cve, date, link])

    wb.save(FILE)

# ================= MAIN =================

def main():
    print("Running BleepingComputer CVE fetch...")

    feed = feedparser.parse(RSS_URL)

    existing = load_existing()
    new_data = set()

    for entry in feed.entries:

        if not hasattr(entry, "published_parsed"):
            continue

        pub = datetime(*entry.published_parsed[:6])
        date_str = pub.strftime("%Y-%m-%d")

        link = entry.link

        print("\nArticle:", entry.title)
        print("Date:", date_str)
        print("Link:", link)

        cves = extract_cves(link)

        if cves:
            print("CVEs:", cves)

        for c in cves:
            if c not in existing:
                new_data.add((c, date_str, link))
                existing.add(c)

    if new_data:
        save_new(new_data)
        print("\nAdded:", new_data)
    else:
        print("\nNo new CVEs")

# ================= RUN =================

if __name__ == "__main__":
    main()
